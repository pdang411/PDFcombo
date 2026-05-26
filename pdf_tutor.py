"""PDF Tutor - Oral Quiz Assistant with Piper TTS + PDF/Excel support"""

# --- Standard library imports ---
import os
import re
import json
import uuid
import difflib
import logging
from io import BytesIO

logging.basicConfig(level=logging.DEBUG, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

try:
    from PIL import Image as PILImage
except Exception:
    PILImage = None

try:
    import pytesseract
except Exception:
    pytesseract = None

# --- Third-party library imports ---
import streamlit as st
import requests
import PyPDF2
import fitz
from openpyxl import load_workbook

# --- Shared pipertalk utilities ---
from pipertalk.tts_utils import pcm_to_wav


def _voice_kwargs() -> dict:
    return dict(
        voice=st.session_state.get("tts_voice"),
        speaker_id=st.session_state.get("tts_speaker_id"),
        length_scale=st.session_state.get("tts_length_scale"),
        noise_scale=st.session_state.get("tts_noise_scale"),
        noise_w_scale=st.session_state.get("tts_noise_w"),
    )


@st.cache_data(ttl=30, show_spinner=False)
def _fetch_voices() -> dict:
    try:
        r = requests.get(f"{TTS_BASE_URL}/voices", timeout=5)
        if r.status_code == 200:
            data = r.json()
            if isinstance(data, dict):
                return data
    except Exception:
        pass
    return {}


@st.cache_data(ttl=5, show_spinner=False)
def _fetch_tts_health() -> dict:
    try:
        r = requests.get(f"{TTS_BASE_URL}/health", timeout=5)
        if r.status_code == 200:
            data = r.json()
            if isinstance(data, dict):
                return data
    except Exception:
        pass
    return {}


# --- Page configuration ---
st.set_page_config(page_title="PDF Tutor", layout="wide")
st.title("PDF Tutor - Oral Quiz Assistant")

# --- Environment / configuration constants ---
TTS_BASE_URL = os.getenv("TTS_BASE_URL", "http://localhost:5000")
PDF_READ_CHUNK = int(os.getenv("PDF_READ_CHUNK_SIZE", str(1024 * 1024)))

# --- Disk cache settings for persisting extracted questions across reruns ---
_CACHE_DIR = "/tmp/teacher_cache"
_MAX_CACHE_AGE = 3600  # seconds; cached files older than this are deleted

TESSERACT_PATH = os.getenv("TESSERACT_PATH", "")


# --- OCR helpers for scanned PDFs ---
def ocr_image_to_text(image_bytes: bytes) -> str:
    if pytesseract is None or PILImage is None:
        return ""
    try:
        if TESSERACT_PATH:
            pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
        img = PILImage.open(BytesIO(image_bytes))
        text = pytesseract.image_to_string(img)
        return text.strip()
    except Exception:
        return ""


def ocr_pdf_page(pdf_bytes: bytes, page_index: int, zoom: float = 2.0) -> str:
    try:
        d = fitz.open(stream=pdf_bytes, filetype="pdf")
        try:
            p = d.load_page(page_index)
            mat = fitz.Matrix(zoom, zoom)
            pix = p.get_pixmap(matrix=mat, alpha=False)
            img_bytes = pix.tobytes("png")
        finally:
            d.close()
        return ocr_image_to_text(img_bytes)
    except Exception:
        return ""


def ocr_pdf_file(pdf_bytes: bytes, zoom: float = 2.0) -> str:
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        try:
            page_count = len(doc)
            if page_count == 0:
                return ""
            results = []
            for i in range(page_count):
                page = doc.load_page(i)
                mat = fitz.Matrix(zoom, zoom)
                pix = page.get_pixmap(matrix=mat, alpha=False)
                img_bytes = pix.tobytes("png")
                text = ocr_image_to_text(img_bytes)
                results.append(text)
        finally:
            doc.close()
        return "\n\n".join(results)
    except Exception:
        return ""


# --- Cache helper: remove stale cache files ---
def _clean_cache():
    now = __import__("time").time()
    try:
        for fname in os.listdir(_CACHE_DIR):
            path = os.path.join(_CACHE_DIR, fname)
            if os.path.isfile(path) and now - os.path.getmtime(path) > _MAX_CACHE_AGE:
                os.remove(path)
    except FileNotFoundError:
        pass


# --- Cache helper: serialise questions to disk, return a cache key ---
def _save_questions_cache(questions):
    os.makedirs(_CACHE_DIR, exist_ok=True)
    _clean_cache()
    ck = uuid.uuid4().hex
    with open(os.path.join(_CACHE_DIR, f"{ck}.json"), "w") as f:
        json.dump(questions, f)
    return ck


# --- Cache helper: load questions by cache key, returns None if missing ---
def _load_questions_cache(ck):
    path = os.path.join(_CACHE_DIR, f"{ck}.json")
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return None


def _save_quiz_state(ck):
    os.makedirs(_CACHE_DIR, exist_ok=True)
    state = {
        "quiz_started": st.session_state.quiz_started,
        "q_index": st.session_state.q_index,
        "answered": st.session_state.answered,
        "speech_processed_at": st.session_state.speech_processed_at,
        "question_added_idx": st.session_state.question_added_idx,
        "results": st.session_state.results,
        "messages": st.session_state.messages,
    }
    with open(os.path.join(_CACHE_DIR, f"{ck}.state"), "w") as f:
        json.dump(state, f)


def _load_quiz_state(ck):
    path = os.path.join(_CACHE_DIR, f"{ck}.state")
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return None


# --- Piper TTS HTTP client: sends synthesis requests to the Piper REST API ---
class PiperHTTPClient:
    def __init__(self, base_url: str, timeout: int = 60):
        self.base_url = base_url.rstrip("/")
        self.timeout = timeout
        self._session = requests.Session()

    def synthesize(
        self,
        text: str,
        voice: str = None,
        speaker_id: int = None,
        length_scale: float = None,
        noise_scale: float = None,
        noise_w_scale: float = None,
    ) -> bytes:
        payload = {"text": text, "response_format": "pcm"}
        if voice:
            payload["voice"] = voice
        if speaker_id is not None:
            payload["speaker_id"] = speaker_id
        if length_scale is not None:
            payload["length_scale"] = length_scale
        if noise_scale is not None:
            payload["noise_scale"] = noise_scale
        if noise_w_scale is not None:
            payload["noise_w_scale"] = noise_w_scale
        resp = self._session.post(
            f"{self.base_url}/synthesize", json=payload, timeout=self.timeout
        )
        resp.raise_for_status()
        sr = int(resp.headers.get("X-Piper-Sample-Rate", 22050))
        return pcm_to_wav(resp.content, sr)


# --- Cached factory: reuse a single PiperHTTPClient across Streamlit reruns ---
@st.cache_resource
def get_piper_client():
    try:
        return PiperHTTPClient(TTS_BASE_URL, timeout=60)
    except Exception as e:
        st.error(f"Piper TTS connection failed: {e}")
        return None


# --- TTS helper: synthesise text and return WAV bytes (or None on failure) ---
def speak_text(text, **_vk):
    client = get_piper_client()
    if not client:
        return None
    try:
        return client.synthesize(text, **_vk)
    except Exception as e:
        st.error(f"TTS failed: {e}")
        return None


# --- File reading helper: chunked read with progress bar ---
def _read_file_bytes(source, chunk_size=PDF_READ_CHUNK):
    """Read file bytes from an UploadedFile or a local path with a progress bar."""
    if isinstance(source, str):
        total = os.path.getsize(source)
        fh = open(source, "rb")
    else:
        total = getattr(source, "size", None)
        fh = source
    buf = BytesIO()
    pbar = st.sidebar.progress(0) if total else None
    read = 0
    while True:
        data = fh.read(chunk_size)
        if not data:
            break
        buf.write(data)
        read += len(data)
        if pbar and total:
            pbar.progress(min(100, int(read / total * 100)))
    if isinstance(source, str):
        fh.close()
    return buf.getvalue()


# --- File parsing: extract (question, answer) pairs from an Excel workbook ---
# Expects Column A = Question, Column B = Answer; skips the header row.
def extract_questions_from_excel(file_bytes):
    wb = load_workbook(BytesIO(file_bytes))
    ws = wb.active
    questions = []
    rows = list(ws.iter_rows(min_col=1, max_col=2, values_only=True))
    for i, row in enumerate(rows):
        if i == 0:
            continue
        q, a = row[0], row[1]
        if q and a and str(q).strip() and str(a).strip():
            questions.append((str(q).strip(), str(a).strip()))
    return questions


# --- File parsing: extract (question, answer) pairs from a PDF ---
# Tries three strategies in order:
#   1. Labelled lines  (Q: / Question: / A: / Answer:)
#   2. Tab-separated pairs
#   3. Lines containing '?' followed by a short answer line
def _parse_qa_lines(lines):
    questions = []
    current_q = None
    for line in lines:
        m_q = re.match(
            r"^(?:Question\s*\d*[.:]|Q\d*[.:])\s*(.*)", line, re.IGNORECASE
        )
        m_a = re.match(
            r"^(?:Answer\s*\d*[.:]|A\d*[.:])\s*(.*)", line, re.IGNORECASE
        )
        if m_q:
            current_q = m_q.group(1)
        elif m_a and current_q:
            questions.append((current_q, m_a.group(1)))
            current_q = None
    if not questions:
        for line in lines:
            if "\t" in line:
                parts = line.split("\t", 1)
                if parts[0].strip() and parts[1].strip():
                    questions.append((parts[0].strip(), parts[1].strip()))
    if not questions:
        for i, line in enumerate(lines):
            if "?" in line and i + 1 < len(lines):
                ans = lines[i + 1]
                if ans and len(ans) < 200:
                    questions.append((line, ans))
    return questions


def _extract_keyword_qa(text):
    """Extract Q/A pairs from text using 'Question:' and 'Answer:' keywords.

    Captures all content after 'Question:' up to '?',
    and all content after 'Answer:' up to the second period
    (or first period if only one exists, or all if none).
    Pairs them in order of appearance.
    """
    q_pat = re.compile(r"Question:\s*(.*?)\?", re.IGNORECASE | re.DOTALL)
    q_matches = [(m.start(), m.group(1).strip()) for m in q_pat.finditer(text)]
    q_matches = [(p, t) for p, t in q_matches if t]

    # Extract answer blocks between "Answer:" and next "Question:" or end
    ref_pat = re.compile(r"\(Reference:[^)]*\)", re.IGNORECASE)
    a_block_pat = re.compile(
        r"Answer:\s*(.*?)(?=Question:|\Z)", re.IGNORECASE | re.DOTALL
    )
    a_matches = []
    for m in a_block_pat.finditer(text):
        block = m.group(1).strip()
        if not block:
            continue
        ref_m = ref_pat.search(block)
        ref_text = ref_m.group(0) if ref_m else ""
        # Strip reference from block before period-trimming
        trimmed = block.replace(ref_text, "", 1).strip() if ref_text else block
        periods = [i for i, ch in enumerate(trimmed) if ch == "."]
        if len(periods) >= 2:
            answer = trimmed[: periods[1] + 1].strip()
        elif len(periods) == 1:
            answer = trimmed[: periods[0] + 1].strip()
        else:
            answer = trimmed
        if ref_text:
            answer += " " + ref_text
        a_matches.append((m.start(), answer))

    pairs = []
    i = j = 0
    while i < len(q_matches) and j < len(a_matches):
        if q_matches[i][0] < a_matches[j][0]:
            pairs.append((q_matches[i][1], a_matches[j][1]))
            i += 1
            j += 1
        else:
            j += 1
    return pairs


def extract_questions_from_pdf(file_bytes, enable_ocr=False):
    text = ""
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        for page in doc:
            blocks = page.get_text("blocks")
            text_blocks = [(b[0], b[1], b[4].strip()) for b in blocks if b[6] == 0 and b[4].strip()]
            if not text_blocks:
                continue
            xs = [b[0] for b in text_blocks]
            mid_x = (min(xs) + max(xs)) / 2
            left = sorted([b for b in text_blocks if b[0] < mid_x], key=lambda b: b[1])
            right = sorted([b for b in text_blocks if b[0] >= mid_x], key=lambda b: b[1])
            for b in left + right:
                text += b[2] + "\n"
        doc.close()
    except Exception:
        try:
            reader = PyPDF2.PdfReader(BytesIO(file_bytes))
            for page in reader.pages:
                text += (page.extract_text() or "") + "\n"
        except Exception:
            return []

    questions = _extract_keyword_qa(text)

    if not questions:
        lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
        questions = _parse_qa_lines(lines)

    if not questions and enable_ocr and pytesseract is not None:
        ocr_text = ocr_pdf_file(file_bytes)
        if ocr_text.strip():
            ocr_lines = [ln.strip() for ln in ocr_text.split("\n") if ln.strip()]
            questions = _parse_qa_lines(ocr_lines)

    return questions


# --- Answer grading: returns True when the user answer is close enough ---
# Strips punctuation, checks for exact match, substring containment,
# then falls back to a SequenceMatcher similarity ratio >= 0.65.
def check_answer(user_answer, correct_answer):
    ua = re.sub(r"[^a-z0-9\s]", "", user_answer.strip().lower()).strip()
    ca = re.sub(r"[^a-z0-9\s]", "", correct_answer.strip().lower()).strip()
    if ua == ca:
        return True
    if ca in ua or ua in ca:
        return True
    ratio = difflib.SequenceMatcher(None, ua, ca).ratio()
    return ratio >= 0.65


# --- Session-state initialisation ---
# All keys used by the quiz are declared here so that every branch of the
# app can assume they already exist in st.session_state.
KEYS = [
    "quiz_started", "q_index", "results", "answered",
    "messages", "question_added_idx", "feedback_audio",
    "speech_processed_at",
]
for k in KEYS:
    if k not in st.session_state:
        if k == "results":
            st.session_state[k] = []
        elif k == "messages":
            st.session_state[k] = []
        elif k == "question_added_idx":
            st.session_state[k] = -1
        elif k == "feedback_audio":
            st.session_state[k] = None
        elif k in ("quiz_started", "answered"):
            st.session_state[k] = False
        elif k == "q_index":
            st.session_state[k] = 0
        elif k == "speech_processed_at":
            st.session_state[k] = -1

# --- Sidebar: TTS status indicator, file uploader, and voice-input widget ---
with st.sidebar:
    st.header("Upload")
    try:
        r = requests.get(f"{TTS_BASE_URL}/health", timeout=3)
        if r.json().get("status") == "ok":
            st.success("Piper TTS connected")
        else:
            st.warning("Piper TTS degraded")
    except Exception:
        st.error("Piper TTS not reachable")

    use_local_file = st.checkbox("Load file from server (faster)")
    local_file_path = None
    if use_local_file:
        pdf_files = [f for f in os.listdir(".") if f.lower().endswith(".pdf")]
        xl_files = [f for f in os.listdir(".") if f.lower().endswith((".xlsx", ".xls"))]
        all_files = sorted(pdf_files + xl_files, key=str.lower)
        if not all_files:
            st.info("No PDF/Excel files found in project folder.")
        else:
            sel = st.selectbox("Choose file from server", all_files)
            local_file_path = os.path.join(os.getcwd(), sel)
            st.caption(f"Selected: {local_file_path}")
        uploaded_file = None
    else:
        uploaded_file = st.file_uploader("Upload PDF or Excel", type=["pdf", "xlsx", "xls"])

    if uploaded_file or local_file_path:
        if st.button("✕ Remove File", use_container_width=True):
            for k in ("questions", "_ck", "quiz_started", "q_index", "results",
                      "answered", "messages", "question_added_idx", "feedback_audio"):
                st.session_state.pop(k, None)
            st.query_params.pop("_ck", None)
            if "questions_file_name" in st.session_state:
                del st.session_state.questions_file_name
            st.rerun()

    if pytesseract is None:
        st.caption(
            "OCR not available (pytesseract missing). Scanned PDFs won't extract text."
        )
    elif TESSERACT_PATH and not os.path.isfile(TESSERACT_PATH):
        st.caption(
            "Tesseract not found at TESSERACT_PATH. Scanned PDFs may fail."
        )

    enable_ocr = st.toggle(
        "OCR for scanned PDFs",
        value=False,
        help="Enable OCR to extract text from scanned/image PDFs.",
    )

    st.divider()
    st.subheader("Voice Settings")
    voices_data = _fetch_voices()
    tts_health = _fetch_tts_health()
    server_model = tts_health.get("model")
    voice_options = ["Server default", *voices_data.keys()]

    # Always reflect backend MODEL after a new voice is downloaded
    if (
        server_model
        and server_model in voices_data
        and st.session_state.get("tts_voice_name") != server_model
    ):
        st.session_state["tts_voice_name"] = server_model
        st.session_state.pop("tts_spk_name", None)
        st.session_state.pop("tts_spk_num", None)

    # If no voice is selected, default to backend MODEL
    if st.session_state.get("tts_voice_name") not in voice_options:
        st.session_state["tts_voice_name"] = (
            server_model if server_model in voices_data else "Server default"
        )

    selected_voice = st.selectbox("TTS voice model", voice_options, key="tts_voice_name")
    speaker_id = None
    if selected_voice != "Server default":
        voice_config = voices_data.get(selected_voice, {})
        num_speakers = int(voice_config.get("num_speakers", 1) or 1)
        if num_speakers > 1:
            speaker_map = voice_config.get("speaker_id_map", {})
            if speaker_map:
                speaker_names = list(speaker_map.keys())
                if st.session_state.get("tts_spk_name") not in speaker_names:
                    st.session_state["tts_spk_name"] = speaker_names[0]
                speaker_name = st.selectbox("Speaker", speaker_names, key="tts_spk_name")
                speaker_id = int(speaker_map[speaker_name])
            else:
                speaker_id = int(
                    st.number_input(
                        "Speaker ID",
                        min_value=0,
                        max_value=max(0, num_speakers - 1),
                        value=0,
                        step=1,
                        key="tts_spk_num",
                    )
                )
    st.session_state["tts_voice"] = None if selected_voice == "Server default" else selected_voice
    st.session_state["tts_speaker_id"] = speaker_id
    st.slider("Length scale", 0.5, 2.0, st.session_state.get("tts_length_scale", 1.0), 0.1, key="tts_length_scale")
    st.slider("Noise scale", 0.0, 1.5, st.session_state.get("tts_noise_scale", 0.667), 0.01, key="tts_noise_scale")
    st.slider("Noise W", 0.0, 1.0, st.session_state.get("tts_noise_w", 0.8), 0.01, key="tts_noise_w")

    st.divider()
    st.subheader("Voice Input")

    with st.form("voice_form", clear_on_submit=True):
        voice_text = st.text_input("Voice command", placeholder="Type a voice command (or use mic)", label_visibility="collapsed")
        st.form_submit_button("Send")

    if voice_text and voice_text.strip():
        st.query_params["speech"] = voice_text.strip()
        st.rerun()

    st.html(
        f"<!-- q{st.session_state.q_index} -->" +
        """
        <style>
            #voice-btn {
                padding: 8px 16px; font-size: 16px; cursor: pointer;
                border-radius: 4px; border: 1px solid #ccc; background: #4CAF50; color: #fff;
                width: 100%;
            }
            #voice-btn:hover { background: #45a049; }
            #voice-btn:disabled { opacity: 0.6; cursor: not-allowed; }
            #voice-btn.listening { background: #22c55e; }
            #voice-status { margin-top: 5px; font-size: 13px; min-height: 20px; }
        </style>
        <button id="voice-btn" type="button">🎤 Speak Answer</button>
        <div id="voice-status"></div>
        <script>
        (function() {
            var SR = window.SpeechRecognition || window.webkitSpeechRecognition;
            var btn = document.getElementById('voice-btn');
            var status = document.getElementById('voice-status');
            if (!SR) {
                status.textContent = 'Not supported. Try Chrome.';
                btn.disabled = true;
                return;
            }
            btn.addEventListener('click', function() {
                if (btn.disabled) return;
                btn.disabled = true;
                btn.className = 'listening';
                btn.textContent = '\u23cd Listening...';
                status.textContent = 'Speak now...';
                var r = new SR();
                r.lang = 'en-US';
                r.continuous = true;
                r.interimResults = false;
                var accumulated = '';
                var silenceTimer = null;
                var SILENCE_DELAY = 4000;
                function submitAnswer(text) {
                    var t = text.trim();
                    if (!t) {
                        status.textContent = 'No speech detected.';
                        btn.disabled = false;
                        btn.className = '';
                        btn.textContent = '\U0001f3a4 Speak Answer';
                        return;
                    }
                    status.textContent = 'Done!';
                    var url = new URL(window.location);
                    url.searchParams.set('speech', t);
                    window.location.search = url.search;
                }
                var maxTimeout = setTimeout(function() {
                    if (silenceTimer) clearTimeout(silenceTimer);
                    try { r.stop(); } catch(e) {}
                    if (accumulated.trim()) { submitAnswer(accumulated); return; }
                    status.textContent = 'Timed out. Try again.';
                    btn.disabled = false;
                    btn.className = '';
                    btn.textContent = '\U0001f3a4 Speak Answer';
                }, 120000);
                r.onresult = function(e) {
                    for (var i = e.resultIndex; i < e.results.length; i++) {
                        if (e.results[i].isFinal) {
                            accumulated += ' ' + e.results[i][0].transcript;
                        }
                    }
                    status.textContent = 'Heard: ' + accumulated.trim();
                    if (silenceTimer) clearTimeout(silenceTimer);
                    silenceTimer = setTimeout(function() {
                        submitAnswer(accumulated);
                    }, SILENCE_DELAY);
                };
                r.onerror = function(e) {
                    clearTimeout(maxTimeout);
                    if (silenceTimer) clearTimeout(silenceTimer);
                    status.textContent = 'Error: ' + e.error;
                    btn.disabled = false;
                    btn.className = '';
                    btn.textContent = '\U0001f3a4 Speak Answer';
                };
                r.onend = function() {
                    clearTimeout(maxTimeout);
                    if (silenceTimer) clearTimeout(silenceTimer);
                    if (accumulated.trim()) { submitAnswer(accumulated); return; }
                    btn.disabled = false;
                    btn.className = '';
                    btn.textContent = '\U0001f3a4 Speak Answer';
                };
                try { r.start(); } catch(err) {
                    status.textContent = 'Error: ' + err.message;
                    btn.disabled = false;
                    btn.className = '';
                    btn.textContent = '\U0001f3a4 Speak Answer';
                    clearTimeout(maxTimeout);
                }
            });
        })();
        </script>
        """,
        unsafe_allow_javascript=True,
    )


# Priority order:
#   1. Already in session_state (page rerun)
#   2. Persisted on disk via _ck query-param (browser refresh / share link)
#   3. Freshly uploaded file
if "questions" not in st.session_state:
    ck = st.query_params.get("_ck")
    if ck:
        cached = _load_questions_cache(ck)
        if cached:
            st.session_state.questions = cached
            st.session_state._ck = ck
            # Restore quiz state from disk after browser reload
            quiz_state = _load_quiz_state(ck)
            if quiz_state:
                for k, v in quiz_state.items():
                    st.session_state[k] = v
if "questions" not in st.session_state:
    source = local_file_path or uploaded_file
    if not source:
        st.info("Upload a PDF or Excel file in the sidebar to begin.")
        st.markdown("""
        ### How it works
        1. Upload a PDF or Excel file with Q&A pairs
        2. Click **Start Quiz** to begin
        3. **Listen** to each question via Piper TTS
        4. Click **🎤 Speak Answer** in the sidebar and say your answer
        5. The answer is **auto-submitted** — feedback appears instantly

        ### File formats
        - **Excel**: Column A = Question, Column B = Answer
        - **PDF**: Q:/A: pairs or tab-separated Q and A lines
        """)
        st.stop()

    file_name = local_file_path or getattr(uploaded_file, "name", "upload")
    ext = os.path.splitext(file_name)[1].lower()
    file_bytes = _read_file_bytes(source)

    with st.spinner("Extracting questions..."):
        if ext in (".xlsx", ".xls"):
            st.session_state.questions = extract_questions_from_excel(file_bytes)
        else:
            st.session_state.questions = extract_questions_from_pdf(file_bytes, enable_ocr=enable_ocr)

    if not st.session_state.questions:
        st.error("No questions found. Use Col A=Q / Col B=A for Excel, or Q:/A: pairs for PDF.")
        st.stop()

    st.session_state.questions_file_name = file_name
    ck = _save_questions_cache(st.session_state.questions)
    st.session_state._ck = ck
    st.query_params["_ck"] = ck
    _save_quiz_state(ck)
    st.sidebar.success(f"Loaded {len(st.session_state.questions)} question(s)")
    with st.sidebar.expander("Debug: Q&A pairs", expanded=False):
        for i, (q, a) in enumerate(st.session_state.questions, 1):
            st.text(f"Q{i}: {q[:60]}...")
            st.text(f"A{i}: {a[:60]}...")
            st.divider()
elif uploaded_file or local_file_path:
    source = local_file_path or uploaded_file
    file_name = local_file_path or getattr(uploaded_file, "name", "upload")
    if st.session_state.get("questions_file_name") == file_name:
        pass  # same file — no re-extraction needed
    else:
        ext = os.path.splitext(file_name)[1].lower()
        file_bytes = _read_file_bytes(source)
        with st.spinner("Re-extracting questions from new file..."):
            if ext in (".xlsx", ".xls"):
                st.session_state.questions = extract_questions_from_excel(file_bytes)
            else:
                st.session_state.questions = extract_questions_from_pdf(file_bytes, enable_ocr=enable_ocr)
        if not st.session_state.questions:
            st.error("No questions found.")
            st.stop()
        st.session_state.questions_file_name = file_name
        ck = _save_questions_cache(st.session_state.questions)
        st.session_state._ck = ck
        st.query_params["_ck"] = ck
        st.sidebar.success(f"Loaded {len(st.session_state.questions)} question(s)")
        with st.sidebar.expander("Debug: Q&A pairs", expanded=False):
            for i, (q, a) in enumerate(st.session_state.questions, 1):
                st.text(f"Q{i}: {q[:60]}...")
                st.text(f"A{i}: {a[:60]}...")
                st.divider()

# --- Convenience aliases for the current question list and position ---
qs = st.session_state.questions
total = len(qs)

log.debug(
    "QUIZ CHECK: quiz_started=%s, q_index=%s, total=%s, has_questions=%s",
    st.session_state.quiz_started,
    st.session_state.q_index,
    total,
    bool(qs),
)

# --- Speech handler ---
speech_text = st.query_params.get("speech")

log.debug(
    "STATE: q_index=%s answered=%s speech_processed_at=%s speech=%s",
    st.session_state.q_index,
    st.session_state.answered,
    st.session_state.speech_processed_at,
    speech_text,
)

# PROCESS SPEECH ONLY ONCE PER QUESTION
if (
    speech_text
    and not st.session_state.answered
    and st.session_state.speech_processed_at != st.session_state.q_index
):

    # RESTORE QUIZ STATE after browser reload from speech JS
    st.session_state.quiz_started = True

    current_idx = st.session_state.q_index

    q_text, a_text = qs[current_idx]

    log.debug(
        "PROCESSING SPEECH FOR QUESTION %s",
        current_idx,
    )

    # PREVENT LOOP IMMEDIATELY
    st.session_state.speech_processed_at = current_idx

    # ENSURE QUESTION EXISTS IN CHAT
    if not st.session_state.messages:

        st.session_state.messages.append(
            {
                "role": "assistant",
                "content": q_text,
            }
        )

    # USER ANSWER
    st.session_state.messages.append(
        {
            "role": "user",
            "content": speech_text,
        }
    )

    # CHECK ANSWER
    correct = check_answer(speech_text, a_text)

    # FEEDBACK
    if correct:
        feedback = f"\u2705 Correct! Answer: {a_text}"
    else:
        feedback = f"\u274c Incorrect. Correct answer: {a_text}"

    # ASSISTANT FEEDBACK
    st.session_state.messages.append(
        {
            "role": "assistant",
            "content": feedback,
        }
    )

    # SAVE RESULT
    st.session_state.results.append(
        {
            "question": q_text,
            "user_answer": speech_text,
            "correct_answer": a_text,
            "is_correct": correct,
        }
    )

    # MARK ANSWERED
    st.session_state.answered = True

    # AUDIO
    fb_audio = speak_text(feedback, **_voice_kwargs())

    if fb_audio:
        st.session_state.feedback_audio = fb_audio

    # REMOVE ONLY SPEECH PARAM
        if "speech" in st.query_params:
            del st.query_params["speech"]

        ck = st.session_state.get("_ck")
        if ck:
            _save_quiz_state(ck)

        st.rerun()


# --- Results screen ---
elif st.session_state.q_index >= total:

    st.balloons()

    st.subheader("Quiz Complete!")

    results = st.session_state.results

    correct = sum(
        1 for r in results
        if r["is_correct"]
    )

    score = round(
        correct / total * 100,
        1,
    ) if total else 0

    st.metric(
        "Score",
        f"{correct} / {total}",
        f"{score}%",
    )

    rows = []

    for i, r in enumerate(results, 1):

        rows.append(
            {
                "#": i,
                "Question": r["question"][:90],
                "Your Answer": r["user_answer"],
                "Correct": r["correct_answer"],
                "Result": (
                    "PASS"
                    if r["is_correct"]
                    else "FAIL"
                ),
            }
        )

    st.dataframe(
        rows,
        use_container_width=True,
    )

    if st.button(
        "New Quiz",
        key="new_q",
        type="primary",
        use_container_width=True,
    ):

        st.session_state.quiz_started = False

        st.session_state.q_index = 0

        st.session_state.results = []

        st.session_state.answered = False

        st.session_state.messages = []

        st.session_state.question_added_idx = -1

        st.session_state.feedback_audio = None

        st.session_state.speech_processed_at = -1

        if "speech" in st.query_params:
            del st.query_params["speech"]

        st.rerun()


# --- Active question screen ---
else:

    current_idx = st.session_state.q_index

    q_text, a_text = qs[current_idx]

    log.debug(
        "ACTIVE QUESTION %s/%s",
        current_idx + 1,
        total,
    )

    st.progress(
        (current_idx + 1) / total,
        text=f"Question {current_idx + 1} of {total}",
    )

    # ADD QUESTION ONLY ONCE
    if st.session_state.question_added_idx != current_idx:

        already_added = any(
            msg["role"] == "assistant"
            and msg["content"] == q_text
            for msg in st.session_state.messages
        )

        if not already_added:

            st.session_state.messages.append(
                {
                    "role": "assistant",
                    "content": q_text,
                }
            )

        st.session_state.question_added_idx = current_idx

    # RENDER CHAT
    for msg in st.session_state.messages:

        with st.chat_message(msg["role"]):

            st.write(msg["content"])

    # PLAY AUDIO
    if st.session_state.feedback_audio:

        st.audio(
            st.session_state.feedback_audio,
            format="audio/wav",
            autoplay=True,
        )

        st.session_state.feedback_audio = None

    # SPEAK QUESTION
    if st.button(
        "Speak Question",
        key="speak_q",
        use_container_width=True,
    ):

        audio = speak_text(q_text, **_voice_kwargs())

        if audio:

            st.audio(
                audio,
                format="audio/wav",
                autoplay=True,
            )

    # NEXT QUESTION
    if st.session_state.answered:

        if st.button(
            "Next Question",
            key=f"next_q_{current_idx}",
            type="primary",
            use_container_width=True,
        ):

            old_idx = st.session_state.q_index

            st.session_state.q_index += 1

            log.debug(
                "NEXT QUESTION: %s -> %s",
                old_idx,
                st.session_state.q_index,
            )

            st.session_state.answered = False

            st.session_state.messages = []

            st.session_state.question_added_idx = -1

            st.session_state.feedback_audio = None

            st.session_state.speech_processed_at = -1

            if "speech" in st.query_params:
                del st.query_params["speech"]

            ck = st.session_state.get("_ck")
            if ck:
                _save_quiz_state(ck)

            st.rerun()

    # SKIP BUTTON
    if not st.session_state.answered:

        cols = st.columns([4, 1])

        with cols[1]:

            if st.button(
                "⏭ Skip",
                key=f"skip_q_{current_idx}",
                use_container_width=True,
            ):

                old_idx = st.session_state.q_index

                st.session_state.q_index += 1

                log.debug(
                    "SKIP QUESTION: %s -> %s",
                    old_idx,
                    st.session_state.q_index,
                )

                st.session_state.answered = False

                st.session_state.messages = []

                st.session_state.question_added_idx = -1

                st.session_state.feedback_audio = None

                st.session_state.speech_processed_at = -1

                if "speech" in st.query_params:
                    del st.query_params["speech"]

                ck = st.session_state.get("_ck")
                if ck:
                    _save_quiz_state(ck)

                st.rerun()

